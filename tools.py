import json
import os
import datetime
import openpyxl
from qfluentwidgets import (
    InfoBarIcon,
    InfoBar as infobar,
    PushButton,
    setTheme,
    Theme,
    FluentIcon,
    InfoBarPosition,
    InfoBarManager,
)
from PyQt5.QtCore import Qt, QRectF


class FileOperation:
    def ReadJSon(file_path):
        with open(file_path, "r") as file:
            data = json.load(file)
        return data

    def WriteJson(file_path, data):
        with open(file_path, "w") as file:
            json.dump(data, file, indent=4)

    def CreateJson(file_path, data):
        if not os.path.exists(os.path.dirname(file_path)):
            try:
                os.makedirs(os.path.dirname(file_path))
            except OSError as e:
                if e.errno != os.errno.EEXIST:
                    raise
        if not os.path.exists(file_path):
            with open(file_path, "w") as file:
                json.dump(data, file)

    def CreateText(file_path, data):
        if not os.path.exists(os.path.dirname(file_path)):
            try:
                os.makedirs(os.path.dirname(file_path))
            except OSError as e:
                if e.errno != os.errno.EEXIST:
                    raise
        if not os.path.exists(file_path):
            with open(file_path, "w") as file:
                file.write(data)

    def WriteText(file_path, data):
        with open(file_path, "w") as file:
            file.write(data)

    def ReadText(file_path):
        with open(file_path, "r") as file:
            return file.context()

    def ReadExcel(path_file):
        workbook = openpyxl.load_workbook(path_file)
        # 获取工作表
        sheet = workbook.active
        data = [[cell.value for cell in row] for row in sheet.iter_rows()]
        workbook.close()
        return data


def WriteExcel(path_file, data):
    """ """
    workbook = openpyxl.load_workbook(path_file)
    sheet = workbook.active
    for row, new_row_data in zip(sheet.iter_rows(), data):
        for cell, new_cell_data in zip(row, new_row_data):
            cell.value = new_cell_data
            workbook.save(path_file)
            workbook.close()


class Logger:
    def Info(data):
        print(f"[INFO][{Logger.GetTime()}] {data}\033[0m ")

    def Warning(data):
        print(f"\033[33m[WARN][{Logger.GetTime()}] {data}\033[0m ")

    def Error(data):
        print(f"\033[31m[ERROR][{Logger.GetTime()}] {data}\033[0m ")

    def Debug(data, Debug_mode):
        """发送一个调试 `Log`
        #### 参数：
            - data `string` 内容
        """
        if Debug_mode:
            print(f"\033[34m[DEBUG][{Logger.GetTime()}] {data}\033[0m ")

    def GetTime():
        """获取当前时间"""
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        return current_time


class InfoBar:
    def Success(self, title, content):
        # convenient class mothod
        infobar.success(
            title,
            content,
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP,
            duration=3000,
            parent=self,
        )

    def Warnning(self, title, content):
        infobar.warning(
            title,
            content,
            orient=Qt.Horizontal,
            isClosable=False,  # disable close button
            position=InfoBarPosition.TOP,
            duration=3000,
            parent=self,
        )

    def Error(self, title, content):
        infobar.error(
            title,
            content,
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP,
            duration=-1,
            parent=self,
        )
